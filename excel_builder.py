"""
Excel workbook generator for PCA Builder.
Produces formatted workbook with 6 tabs from PCA source data.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="1B1464")
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
CURRENCY = '$#,##0'
DECIMAL = '$#,##0.00'
NUMBER = '#,##0'
PCT = '0.0%'
PCT2 = '0.00%'
thin = Side(style="thin", color="DDDDDD")
BORDER = Border(top=thin, bottom=thin, left=thin, right=thin)


def _write_headers(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER


def _auto_width(ws):
    for col in range(1, ws.max_column + 1):
        max_len = 0
        for row in range(1, min(ws.max_row + 1, 50)):
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 10), 25)


def build_pca_workbook(source_data: dict, output_path: str) -> str:
    wb = Workbook()

    # ── Tab 1: Campaign Overview ──
    ws = wb.active
    ws.title = "1. Overview"
    headers = ["Platform", "Spend", "Impressions", "Clicks", "Completed Views", "CPC", "CPM", "CPCV"]
    _write_headers(ws, headers)
    fmts = [None, CURRENCY, NUMBER, NUMBER, NUMBER, DECIMAL, DECIMAL, DECIMAL]

    for i, row in enumerate(source_data["overview"], 2):
        ws.cell(row=i, column=1, value=row.get("platform", "").replace("_", " ").title()).border = BORDER
        ws.cell(row=i, column=2, value=row.get("total_spend", 0)).border = BORDER
        ws.cell(row=i, column=3, value=row.get("total_impressions", 0)).border = BORDER
        ws.cell(row=i, column=4, value=row.get("total_clicks", 0)).border = BORDER
        ws.cell(row=i, column=5, value=row.get("total_completed_views", 0)).border = BORDER
        ws.cell(row=i, column=6, value=row.get("cpc", "")).border = BORDER
        ws.cell(row=i, column=7, value=row.get("cpm", "")).border = BORDER
        ws.cell(row=i, column=8, value=row.get("cpcv", "")).border = BORDER

    for col, fmt in enumerate(fmts, 1):
        if fmt:
            for row in range(2, len(source_data["overview"]) + 2):
                ws.cell(row=row, column=col).number_format = fmt

    total_row = len(source_data["overview"]) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
    for col in [2, 3, 4, 5]:
        letter = get_column_letter(col)
        ws.cell(row=total_row, column=col, value=f"=SUM({letter}2:{letter}{total_row-1})")
        ws.cell(row=total_row, column=col).font = BOLD_FONT
        ws.cell(row=total_row, column=col).number_format = fmts[col-1]

    ws.auto_filter.ref = f"A1:H{total_row - 1}"
    ws.freeze_panes = "A2"
    _auto_width(ws)

    # ── Tab 2: Weekly Trends ──
    ws2 = wb.create_sheet("2. Weekly Trends")
    headers2 = ["Week", "Platform", "Spend", "Impressions", "CPM", "Prev Week Spend", "Spend WoW %"]
    _write_headers(ws2, headers2)
    fmts2 = [None, None, CURRENCY, NUMBER, DECIMAL, CURRENCY, '0.0"%"']

    for i, row in enumerate(source_data["weekly_trends"], 2):
        ws2.cell(row=i, column=1, value=row.get("week_start", "")).border = BORDER
        ws2.cell(row=i, column=2, value=row.get("platform", "").replace("_", " ").title()).border = BORDER
        ws2.cell(row=i, column=3, value=row.get("spend", 0)).border = BORDER
        ws2.cell(row=i, column=4, value=row.get("impressions", 0)).border = BORDER
        ws2.cell(row=i, column=5, value=row.get("cpm", 0)).border = BORDER
        ws2.cell(row=i, column=6, value=row.get("prev_spend", "")).border = BORDER
        ws2.cell(row=i, column=7, value=row.get("spend_wow_pct", "")).border = BORDER

    for col, fmt in enumerate(fmts2, 1):
        if fmt:
            for row in range(2, len(source_data["weekly_trends"]) + 2):
                ws2.cell(row=row, column=col).number_format = fmt

    ws2.auto_filter.ref = f"A1:G{len(source_data['weekly_trends']) + 1}"
    ws2.freeze_panes = "A2"
    _auto_width(ws2)

    # ── Tab 3: Breakdowns ──
    ws3 = wb.create_sheet("3. Breakdowns")
    current_row = 1
    for dim_name, dim_data in source_data["breakdowns"].items():
        label = dim_name.replace("by_", "").replace("_", " ").title()
        ws3.cell(row=current_row, column=1, value=f"BY {label.upper()}")
        ws3.cell(row=current_row, column=1).font = Font(name="Arial", bold=True, size=12, color="1B1464")
        current_row += 1

        dim_headers = [label, "Spend", "Impressions", "CPM", "CPC", "CPCV", "Spend Share %"]
        _write_headers(ws3, dim_headers, row=current_row)
        current_row += 1

        for row in dim_data:
            ws3.cell(row=current_row, column=1, value=row.get("value", "")).border = BORDER
            ws3.cell(row=current_row, column=2, value=row.get("spend", 0)).border = BORDER
            ws3.cell(row=current_row, column=2).number_format = CURRENCY
            ws3.cell(row=current_row, column=3, value=row.get("impressions", 0)).border = BORDER
            ws3.cell(row=current_row, column=3).number_format = NUMBER
            ws3.cell(row=current_row, column=4, value=row.get("cpm", 0)).border = BORDER
            ws3.cell(row=current_row, column=4).number_format = DECIMAL
            ws3.cell(row=current_row, column=5, value=row.get("cpc", "")).border = BORDER
            ws3.cell(row=current_row, column=5).number_format = DECIMAL
            ws3.cell(row=current_row, column=6, value=row.get("cpcv", "")).border = BORDER
            ws3.cell(row=current_row, column=6).number_format = DECIMAL
            ws3.cell(row=current_row, column=7, value=row.get("spend_share_pct", 0)).border = BORDER
            ws3.cell(row=current_row, column=7).number_format = '0.0"%"'
            current_row += 1
        current_row += 1

    _auto_width(ws3)

    # ── Tab 4: Benchmarks ──
    ws4 = wb.create_sheet("4. Benchmarks")
    headers4 = ["Platform", "Metric", "Actual", "Benchmark", "Variance %", "Type", "Sample Size"]
    _write_headers(ws4, headers4)

    for i, row in enumerate(source_data["benchmarks"], 2):
        ws4.cell(row=i, column=1, value=row.get("platform", "").replace("_", " ").title()).border = BORDER
        ws4.cell(row=i, column=2, value=row.get("metric_name", "")).border = BORDER
        ws4.cell(row=i, column=3, value=row.get("actual_value", 0)).border = BORDER
        ws4.cell(row=i, column=3).number_format = DECIMAL
        ws4.cell(row=i, column=4, value=row.get("benchmark_value", 0)).border = BORDER
        ws4.cell(row=i, column=4).number_format = DECIMAL
        ws4.cell(row=i, column=5, value=row.get("variance_pct", 0)).border = BORDER
        ws4.cell(row=i, column=5).number_format = '0.0"%"'
        ws4.cell(row=i, column=6, value=row.get("benchmark_type", "")).border = BORDER
        ws4.cell(row=i, column=7, value=row.get("sample_size", "")).border = BORDER

    ws4.auto_filter.ref = f"A1:G{len(source_data['benchmarks']) + 1}"
    ws4.freeze_panes = "A2"
    _auto_width(ws4)

    # ── Tab 5: Plan vs Actual ──
    ws5 = wb.create_sheet("5. Plan vs Actual")
    if source_data.get("plan_vs_actual"):
        headers5 = ["Platform", "Planned Spend", "Actual Spend", "Variance %"]
        _write_headers(ws5, headers5)
        for i, row in enumerate(source_data["plan_vs_actual"], 2):
            ws5.cell(row=i, column=1, value=row.get("platform", "").replace("_", " ").title()).border = BORDER
            ws5.cell(row=i, column=2, value=row.get("planned_spend", 0)).border = BORDER
            ws5.cell(row=i, column=2).number_format = CURRENCY
            ws5.cell(row=i, column=3, value=row.get("actual_spend", 0)).border = BORDER
            ws5.cell(row=i, column=3).number_format = CURRENCY
            ws5.cell(row=i, column=4, value=row.get("spend_variance_pct", 0)).border = BORDER
            ws5.cell(row=i, column=4).number_format = '0.0"%"'
        ws5.auto_filter.ref = f"A1:D{len(source_data['plan_vs_actual']) + 1}"
        ws5.freeze_panes = "A2"
    else:
        ws5.cell(row=1, column=1, value="No media plan data available for this campaign")
        # FIXED: Changed italics=True to italic=True
        ws5.cell(row=1, column=1).font = Font(name="Arial", italic=True, color="888888")
    _auto_width(ws5)

    # ── Tab 6: Raw Campaign Data ──
    ws6 = wb.create_sheet("6. Raw Data")
    raw_headers = ["Date", "Platform", "Campaign", "Tactic", "Targeting", "Device",
                   "Creative Format", "State", "Spend", "Impressions", "Clicks",
                   "Completed Views", "CPC", "CPM", "CPCV", "CTR %"]
    _write_headers(ws6, raw_headers)
    raw_fmts = {9: CURRENCY, 10: NUMBER, 11: NUMBER, 12: NUMBER,
                13: DECIMAL, 14: DECIMAL, 15: DECIMAL, 16: '0.00"%"'}

    raw_data = source_data.get("raw_data", [])
    for i, row in enumerate(raw_data[:10000], 2):
        keys = ["date", "platform", "campaign_name", "tactic", "targeting", "device",
                "creative_format", "state", "spend", "impressions", "clicks",
                "completed_views", "cpc", "cpm", "cpcv", "ctr"]
        for j, key in enumerate(keys, 1):
            c = ws6.cell(row=i, column=j, value=row.get(key, ""))
            c.border = BORDER
            if j in raw_fmts:
                c.number_format = raw_fmts[j]

    ws6.auto_filter.ref = f"A1:{get_column_letter(len(raw_headers))}{min(len(raw_data) + 1, 10001)}"
    ws6.freeze_panes = "A2"
    _auto_width(ws6)

    wb.save(output_path)
    return output_path


# ─── Accent fills for media strategy workbook ───────────────────────────────
_NAVY_FILL   = PatternFill("solid", fgColor="1B1464")
_ORANGE_FILL = PatternFill("solid", fgColor="F05C2C")
_PURPLE_FILL = PatternFill("solid", fgColor="4338CA")
_GREY_FILL   = PatternFill("solid", fgColor="F4F5F7")

_SECTION_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
_SUBHEAD_FONT  = Font(name="Arial", bold=True, color="1B1464", size=10)
_BODY_FONT2    = Font(name="Arial", size=10)
_BODY_BOLD     = Font(name="Arial", bold=True, size=10)

